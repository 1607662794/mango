'''该文件用于从原始excel中提取信息，构建json文件'''
import os
import json
import pandas as pd
from openpyxl import load_workbook
from docx import Document

# 定义文件路径
excel_file_path = '导览语音对照表_填充后.xlsx'
folders = {
    '湖南人': '湖南人-doc',
    '马王堆': '马王堆-doc'
}

# 加载 Excel 文件
workbook = load_workbook(excel_file_path)

# 初始化 JSON 列表
json_data = []


# 定义函数读取 Word 文件内容
def read_word_content(file_path):
    doc = Document(file_path)
    content = []
    for para in doc.paragraphs:
        content.append(para.text)
    full_content = '\n'.join(content)
    # 提取第二个换行符后的内容
    split_content = full_content.split('\n', 2)
    if len(split_content) > 2:
        return split_content[2]
    else:
        return full_content


# 遍历工作簿和工作表
for sheet_name in workbook.sheetnames:
    ws = workbook[sheet_name]

    audio_dict = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        audio_name = row[0].value
        mentioned_cultural_relic = row[1].value
        formal_cultural_relic_name = row[2].value
        cultural_relic_id = row[3].value

        if audio_name:
            # 去掉两边的空格
            mentioned_cultural_relic = mentioned_cultural_relic.strip() if mentioned_cultural_relic else None
            formal_cultural_relic_name = formal_cultural_relic_name.strip() if formal_cultural_relic_name else None
            cultural_relic_id = cultural_relic_id.strip() if cultural_relic_id else None

            # 查找对应的文件夹和文件
            folder_path = folders[sheet_name]
            subfolder_path = os.path.join(folder_path, str(audio_name))
            word_file_path = os.path.join(subfolder_path, f"{audio_name}_原文.docx")

            if os.path.exists(word_file_path):
                file_content = read_word_content(word_file_path)

                if audio_name not in audio_dict:
                    audio_dict[audio_name] = {
                        "file_path": word_file_path,
                        "relate_cultural_relic_name": [],
                        "file_content": file_content,
                        "classification": sheet_name
                    }

                audio_dict[audio_name]["relate_cultural_relic_name"].append({
                    "document_entity_name": mentioned_cultural_relic,
                    "formal_entity_name": formal_cultural_relic_name,
                    "cultural_relic_id": cultural_relic_id
                })

    for audio_name, data in audio_dict.items():
        json_data.append(data)

# 保存 JSON 数据到文件
json_file_path = '导览语音对照表.json'
with open(json_file_path, 'w', encoding='utf-8') as f:
    json.dump(json_data, f, ensure_ascii=False, indent=4)

print(f"JSON 文件已保存为 '{json_file_path}'")
