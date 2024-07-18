'''该文档用于填充原excel文档，便于后续处理。'''
from openpyxl import load_workbook


# 加载 Excel 文件
file_path = '导览语音对照表(1).xlsx'
workbook = load_workbook(file_path)


# 定义函数用于填充空格
def fill_audio_names(sheet):
    ws = workbook[sheet]

    # 定义列索引，假设表头在第一行
    audio_col = 1  # 音频名称
    mention_col = 4  # 提到文物

    last_valid_audio_name = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        audio_cell = row[audio_col - 1]
        mention_cell = row[mention_col - 1]

        if audio_cell.value is None:
            if mention_cell.value is not None:
                audio_cell.value = last_valid_audio_name
        else:
            last_valid_audio_name = audio_cell.value


# 填充两个工作簿中的音频名称列
sheets = ['马王堆', '湖南人']

for sheet in sheets:
    fill_audio_names(sheet)

# 保存结果到新的 Excel 文件
output_file_path = '导览语音对照表_填充后.xlsx'
workbook.save(output_file_path)

print(f"填充完成并保存为 '{output_file_path}'")
